# import module
import datetime
import openpyxl

employer_names = ["وزارت نیرو",
                  "وزارت آموزش و پرورش",
                  "سازمان ثبت اسناد و املاک کشور",
                  "وزارت ارتباطات و فناوری اطلاعات",
                  "وزارت امور اقتصاد و دارایی",
                  "وزارت بهداشت، درمان و آموزش پزشکی",
                  "وزارت جهاد کشاورزی",
                  "وزارت راه و شهرسازی",
                  "وزارت علوم، تحقیقات و فناوری",
                  "وزارت فرهنگ و ارشاد اسلامی",
                  "وزارت کشور",
                  "وزارت میراث فرهنگی، گردشگری و صنایع دستی",
                  "سازمان زمین‌شناسی",
                  "سازمان ملی استاندارد",
                  "مرکز آمار ایران",
                  "سازمان نقشه‌برداری ایران",
                  "مرکز ملی رقابت",
                  "وزارت دادگستری",
                  "معاونت حقوقی ریاست جمهوری",
                  "سازمان برنامه و بودجه کشور",
                  "وزارت نفت",
                  "وزارت دفاع و پشتیبانی نیروهای مسلح",
                  "قوه قضائیه",
                  "نهاد ریاست جمهوری",
                  "سازمان انرژی اتمی ایران",
                  "وزارت صنعت، معدن و تجارت",
                  "وزارت ورزش و جوانان",
                  "وزارت تعاون،کار و رفاه اجتماعی",
                  "سازمان ملی استاندارد ایران",
                  "سازمان اداری و استخدامی کشور"
                  ]

# load excel with its path
wrkbk = openpyxl.load_workbook("27dey.xlsx")

sh = wrkbk.active
data = ''

sql_template = """INSERT INTO `#inojobs` (`id`, `group`, `employer_id`, `emp`, `status`, `title`, `description`, `image`, `location`, `degree`, `pin`, `major`, `date`, `count`, `facilities`, `properties`, `created_at`, `updated_at`) VALUES {data};"""

item_pk = ''

def replace_arabic_char(value):
    try:
        value =  value.replace('ي', 'ی').replace('ك', 'ک').strip()
        if value == 'آموزش و پرورش':
            value = "وزارت آموزش و پرورش"
        if value == "وزارت  علوم تحقیقات و فناوری" or value == "وزارت علوم تحقیقات و فناوری":
            value = "وزارت علوم، تحقیقات و فناوری"
        return value
    except:
        return value


# iterate through excel and display data
for i in range(1, sh.max_row + 1):
    item_pk = sh.cell(row=i, column=1).value
    employer_id = employer_names.index(replace_arabic_char(sh.cell(row=i, column=2).value)) + 1
    employer_subcat_name = replace_arabic_char(sh.cell(row=i, column=3).value)
    job_title = replace_arabic_char(sh.cell(row=i, column=4).value)
    city = replace_arabic_char(sh.cell(row=i, column=6).value)
    province = replace_arabic_char(sh.cell(row=i, column=5).value)
    degree = replace_arabic_char(sh.cell(row=i, column=7).value)
    total_count = replace_arabic_char(sh.cell(row=i, column=8).value)
    free_woman = replace_arabic_char(sh.cell(row=i, column=9).value)
    free_man = replace_arabic_char(sh.cell(row=i, column=10).value)
    free_both = replace_arabic_char(sh.cell(row=i, column=11).value)
    isar_25_woman = replace_arabic_char(sh.cell(row=i, column=12).value)
    isar_25_man = replace_arabic_char(sh.cell(row=i, column=13).value)
    isar_25_both = replace_arabic_char(sh.cell(row=i, column=14).value)
    isar_5_woman = replace_arabic_char(sh.cell(row=i, column=15).value)
    isar_5_man = replace_arabic_char(sh.cell(row=i, column=16).value)
    isar_5_both = replace_arabic_char(sh.cell(row=i, column=17).value)
    malool_woman = replace_arabic_char(sh.cell(row=i, column=17).value)
    malool_man = replace_arabic_char(sh.cell(row=i, column=18).value)
    malool_both = replace_arabic_char(sh.cell(row=i, column=19).value)
    row_query_string = f"""({item_pk}, '{{"id":1,"cat":1}}', {employer_id}, '{{"id": 1, "name": "{employer_subcat_name}"}}', 0, '{job_title}', NULL, '', '{{"spot": "{city}", "province": "{province}"}}', '{degree}', '[]', '["all"]', '{{"publishDate": null,"endDate" : null}}', '{{"count": {total_count},"likes": [], "views": [], "requests": 0, "free_man": {free_man}, "free_woman": {free_woman}, "free_both": {free_both}, "isar_5_man": {isar_5_man}, "isar_5_woman": {isar_5_woman}, "isar_5_both": {isar_5_both}, "isar_25_man": {isar_25_man}, "isar_25_woman": {isar_25_woman}, "isar_25_both": {isar_25_both},"malool_man": {malool_man}, "malool_woman": {malool_woman}, "malool_both": {malool_both}}}', '[]', '{{"type":"gf", "site":null, "salary":null, "gender" : null,"age" : null}}', '{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}', '{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}'),"""
    row_query_string = row_query_string.replace('None', '0')
    data += row_query_string

final_query = sql_template.format(data=data[:-1])
print(final_query)
with open("query.sql", 'w', encoding='utf-8') as f:
    f.write(final_query)

